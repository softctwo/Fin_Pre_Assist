"""文档处理服务"""

import os
import docx
from pypdf import PdfReader
import openpyxl


class DocumentProcessor:
    """文档处理器"""

    @staticmethod
    def extract_text_from_docx(file_path: str) -> str:
        """从Word文档提取文本"""
        try:
            doc = docx.Document(file_path)
            text = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text.append(paragraph.text)
            return "\n".join(text)
        except Exception as e:
            raise Exception(f"读取Word文档失败: {str(e)}")

    @staticmethod
    def extract_text_from_pdf(file_path: str) -> str:
        """从PDF文档提取文本"""
        try:
            text = []
            with open(file_path, "rb") as file:
                pdf_reader = PdfReader(file)
                for page in pdf_reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text.append(page_text)
            return "\n".join(text)
        except Exception as e:
            raise Exception(f"读取PDF文档失败: {str(e)}")

    @staticmethod
    def extract_text_from_xlsx(file_path: str) -> str:
        """从Excel文档提取文本"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text.append(f"=== {sheet_name} ===")
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text.append(row_text)
            return "\n".join(text)
        except Exception as e:
            raise Exception(f"读取Excel文档失败: {str(e)}")

    @staticmethod
    def extract_text_from_txt(file_path: str) -> str:
        """从文本文件提取文本"""
        # 尝试多种编码
        encodings = ["utf-8", "gbk", "gb2312", "latin-1", "iso-8859-1"]

        for encoding in encodings:
            try:
                with open(file_path, "r", encoding=encoding) as file:
                    return file.read()
            except (UnicodeDecodeError, LookupError):
                continue

        # 如果所有编码都失败，尝试二进制模式并忽略错误
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as file:
                return file.read()
        except Exception as e:
            raise Exception(f"读取文本文件失败: {str(e)}")

    @classmethod
    def extract_text(cls, file_path: str) -> str:
        """根据文件类型提取文本"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        file_ext = os.path.splitext(file_path)[1].lower()

        if file_ext in [".doc", ".docx"]:
            return cls.extract_text_from_docx(file_path)
        elif file_ext == ".pdf":
            return cls.extract_text_from_pdf(file_path)
        elif file_ext in [".xls", ".xlsx"]:
            return cls.extract_text_from_xlsx(file_path)
        elif file_ext == ".txt":
            return cls.extract_text_from_txt(file_path)
        else:
            raise ValueError(f"不支持的文件类型: {file_ext}")
